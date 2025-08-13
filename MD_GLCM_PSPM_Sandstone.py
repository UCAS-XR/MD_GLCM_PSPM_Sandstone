import cv2
import numpy as np
from pathlib import Path
from skimage.feature import graycomatrix, graycoprops
from skimage import color
from openpyxl import Workbook
import matplotlib.pyplot as plt


def rgb2gray(img_0, root_path=None, img_path=None):
    if len(img_0.shape) == 1:
        img_gray = img_0
    elif len(img_0.shape) == 3:
        img_0 = cv2.cvtColor(img_0, cv2.COLOR_BGR2RGB)
        img_gray = color.rgb2gray(img_0) * 255
    else:
        raise ValueError("ValueError: Number of image channels not supported!")
    img_gray = img_gray.astype(np.uint8)
    output_path = Path(root_path) / f"{Path(img_path).stem}.jpg"
    save_dir = output_path.parent
    save_dir.mkdir(parents=True, exist_ok=True)
    cv2.imwrite(output_path, img_gray)
    print("Gray-scale image saved successfully!")
    print("save path:", output_path)
    return img_gray


def remove_reflections(img_gray, root_path=None, img_path=None, kernel_size=3, iterations=1, inpaintRadius=3, flags=cv2.INPAINT_NS):
    values, counts = np.unique(img_gray, return_counts=True)
    max_idx = np.argmax(counts)
    print("the most dominant gray-scale value (m) = ", values[max_idx], "counts (M) = ", counts[max_idx])
    valid_indices = np.where((values >= values[max_idx]) & (values <= 255) & (counts <= 0.001 * counts[max_idx]))[0]
    if len(valid_indices) > 0:
        print("G_threshold = ", values[valid_indices[0]], "G_threshold counts = ", counts[valid_indices[0]])
        binary_mask = (img_gray > values[valid_indices[0]]).astype(np.uint8)
        mask_path = Path(root_path) / f"{Path(img_path).stem}.jpg"
        save_dir = mask_path.parent
        save_dir.mkdir(parents=True, exist_ok=True)
        cv2.imwrite(mask_path, binary_mask*255)
        print("binary mask saved successfully!")
        print("save path:", mask_path)
        kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (kernel_size, kernel_size))
        dilated_mask = cv2.dilate(binary_mask, kernel, iterations=iterations)
        img_repaired = cv2.inpaint(img_gray, dilated_mask, inpaintRadius=inpaintRadius, flags=flags)
    else:
        img_repaired = img_gray
    repaired_path = Path(root_path) / "img_repaired" / f"{Path(img_path).stem}.jpg"
    save_dir = repaired_path.parent
    save_dir.mkdir(parents=True, exist_ok=True)
    cv2.imwrite(repaired_path, img_repaired)
    print("Repaired image saved successfully!")
    print("save path:", repaired_path)
    return img_repaired


def histogram_stretching(img_gray, root_path=None, img_path=None):
    random_noise = np.random.rand(*img_gray.shape) - 0.5
    img_float = np.clip(img_gray + random_noise, 0, 255).astype(np.float32)
    min_val, max_val = np.min(img_float), np.max(img_float)
    if min_val == 0 and max_val == 255:
        img_256 = img_float
    else:
        img_256 = (img_float - min_val) * (255.0 / (max_val - min_val))
    output_path = Path(root_path) / f"{Path(img_path).stem}.jpg"
    save_dir = output_path.parent
    save_dir.mkdir(parents=True, exist_ok=True)
    cv2.imwrite(output_path, img_256)
    print("Stretched image saved successfully!")
    print("save path:", output_path)
    return img_256


def compress_grayscale(img_256, levels, root_path=None, img_path=None):
    global img_comp
    valid_levels = {2, 4, 8, 16, 32, 64, 128}
    if levels == 256:
        img_comp = img_256.astype(np.uint8)
    elif levels in valid_levels:
        scale = 256 // levels
        img_comp = ((img_256 // scale) * scale).astype(np.uint8)
    else:
        raise ValueError("ValueError:Invalid gray-scale compression level")

    output_path = Path(root_path) / f"{Path(img_path).stem}_{levels}.jpg"
    save_dir = output_path.parent
    save_dir.mkdir(parents=True, exist_ok=True)
    cv2.imwrite(output_path, img_comp)
    print(f"{levels} level gray-scale compressed image saved successfully!")
    print("save path:", output_path)
    return img_comp


def hist_diagram(img_comp, root_path=None, img_path=None):
    plt.figure(figsize=(12, 6))
    plt.hist(img_comp.ravel(),
             bins=np.max(img_comp),
             range=(0, np.max(img_comp)),
             color='steelblue',
             edgecolor='black',
             alpha=0.7)
    plt.title('Grayscale Image Pixel Distribution', fontsize=14, pad=20)
    plt.xlabel('Pixel Intensity (0-255)', fontsize=12)
    plt.ylabel('Number of Pixels', fontsize=12)
    plt.grid(axis='y', linestyle='--', alpha=0.5)
    output_path = Path(root_path) / "results" / "hist_diagram" / f"hist_diagram_{Path(img_path).stem}.jpg"
    save_dir = output_path.parent
    save_dir.mkdir(parents=True, exist_ok=True)
    plt.savefig(output_path)
    print("Gray-scale histogram saved successfully!")
    print("save path:", output_path)
    plt.close()


def glcm_texture_features(img_comp, properties, distances, angles, levels=None):
    glcm_matrix = graycomatrix(img_comp,
                        distances=distances,
                        angles=angles,
                        levels=levels,
                        symmetric=True,
                        normed=True)
    features = {}
    for prop in properties:
        features[prop] = graycoprops(glcm_matrix, prop)
    return features, glcm_matrix


def save_glcm_matrix(glcm_matrix, root_path=None, img_path=None):
    angle_dic = {0: "0°", 1: "45°", 2: "90°", 3: "135°"}
    for i in range(0, glcm_matrix.shape[2]):
        for j in range(0, glcm_matrix.shape[3]):
            glcm_path = Path(root_path) / f"{Path(img_path).stem}_{i}_{j}.jpg"
            save_dir = glcm_path.parent
            save_dir.mkdir(parents=True, exist_ok=True)
            plt.figure(figsize=(6, 6))
            plt.imshow(glcm_matrix[:, :, i, j], cmap='jet', interpolation='nearest')
            plt.colorbar()
            plt.title(f'GLCM (distance={i+1}, angle={angle_dic[j]}')
            plt.xlabel('Gray Level')
            plt.ylabel('Gray Level')
            plt.savefig(glcm_path)
            print("GLCM matrix saved successfully!")
            print("save path:", glcm_path)
            plt.clf()


def visualize_features(features, distances, root_path=None, img_path=None):
    properties = list(features.keys())
    plt.figure(figsize=(15, 10))
    wb = Workbook()
    ws = wb.active
    title_list = ["0°", "45°", "90°", "135°"]
    for col_idx, value in enumerate(title_list, start=2):
        ws.cell(row=1, column=col_idx, value=value)
    for row_idx, value in enumerate(properties, start=0):
        ws.cell(row=(len(distances) + 1) * row_idx + 2, column=1, value=value)
    for i, prop in enumerate(properties):
        data = features[prop]
        for j in range(data.shape[0]):
            for k in range(data.shape[1]):
                ws.cell(row=(len(distances) + 1) * i + j + 2, column=k + 2, value=data[j, k])
    xlsx_path = Path(root_path) / f"{Path(img_path).stem}.xlsx"
    save_dir = xlsx_path.parent
    save_dir.mkdir(parents=True, exist_ok=True)
    wb.save(xlsx_path)
    print("GLCM features saved successfully!")
    print("save path:", xlsx_path)


if __name__ == "__main__":
    root_path = r"D:\GLCM"
    levels = 256
    distances = list(range(1, 301))
    angles = [0, np.pi / 4, np.pi / 2, 3 * np.pi / 4]
    properties = ['contrast', 'ASM', 'entropy', 'homogeneity', 'dissimilarity', 'correlation']
    extensions = ['.bmp', '.png', '.jpg', 'jpeg']
    image_path = Path(root_path) / "img_RGB"
    filename_list = [file.name for file in image_path.rglob("*")
                     if file.is_file() and file.suffix.lower() in extensions]
    print("filename list:", filename_list)
    for item in filename_list:
        img_path = Path(image_path / item)
        img_0 = cv2.imread(img_path)
        img_gray = rgb2gray(img_0)
        img_repaired = remove_reflections(img_gray, kernel_size=3, iterations=1, inpaintRadius=3, flags=cv2.INPAINT_NS)
        img_256 = histogram_stretching(img_repaired)
        img_comp = compress_grayscale(img_256, levels).astype(np.uint8)
        hist_diagram(img_comp)
        features, glcm = glcm_texture_features(img_comp,
                                               properties=properties,
                                               distances=distances,
                                               angles=angles
                                               )
        for key, val in features.items():
            print(f"{key}: {val.shape}")
        visualize_features(features, distances)
        save_glcm_matrix(glcm)
