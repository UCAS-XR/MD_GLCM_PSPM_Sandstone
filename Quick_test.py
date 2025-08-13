import os
import unittest
import shutil
from PIL import Image
import cv2
import numpy as np
from pathlib import Path
from openpyxl import load_workbook


class test_center_crop(unittest.TestCase):  # Test_1
    def setUp(self):
        self.root = Path(__file__).parent
        self.input_file = self.root / "test_image.bmp"
        self.crop_path = self.root / "results" / "img_crop"
        self.crop_path.mkdir(parents=True, exist_ok=True)

        if self.input_file.exists():
            print("\n✅ Test image import completed!")
        else:
            raise FileNotFoundError("\n❌ Test image does not exist!")

        with Image.open(self.input_file) as img:
            self.input_width, self.input_height = img.size
            self.input_mode = img.mode
            self.input_channels = len(img.getbands())

            print("\n" + "=" * 50)
            print("Test information:")
            print(f"test_image: {self.input_file}")
            print(f"image_size: {self.input_width}x{self.input_height}")
            print(f"image_color_mode: {self.input_mode}")
            print(f"image_channel: {self.input_channels}")
            print("=" * 50 + "\n")

    def test_center_crop(self):
        from Center_crop import center_crop
        center_crop(self.root, self.crop_path)
        print("\n✅ Center crop completed!")
        self.crop_file = self.crop_path / "test_image.bmp"
        if self.crop_file.exists():
            print("\n✅ Img_crop saved!")
        else:
            raise FileNotFoundError("\n❌ Test image center crop failed!")

        with Image.open(self.crop_file) as img:
            width, height = img.size
            self.assertEqual(width, 3072, f"\n❌ Image width should be 3072, got {width}")
            self.assertEqual(height, 3072, f"\n❌ Image height should be 3072, got {height}")
            print(f"✅ Image size meets requirement of 3072*3072 pixels")
            print(f"✅✅ Test_1 verification passed!")


class test_image_processing(unittest.TestCase):  # Test_2
    def setUp(self):
        self.levels = 256
        self.distances = [1, 2, 3]
        self.angles = [0, np.pi / 4, np.pi / 2, 3 * np.pi / 4]
        self.properties = ['contrast', 'ASM', 'entropy']

        self.root = Path(__file__).parent
        self.crop_file = self.root / "results" / "img_crop" / "test_image.bmp"
        self.results_path = self.root / "results"
        self.dirs = {
            "img_gray": self.results_path / "img_gray",
            "binary_mask": self.results_path / "binary_mask",
            "img_repaired": self.results_path / "img_repaired",
            "img_256": self.results_path / "img_256",
            "img_comp": self.results_path / "img_comp",
            "GLCM_matrix": self.results_path / "GLCM_matrix",
            "GLCM_feature_values": self.results_path / "GLCM_feature_values"
        }

        for dir_path in self.dirs.values():
            dir_path.mkdir(parents=True, exist_ok=True)

    def test_image_processing(self):
        from MD_GLCM_PSPM_Sandstone import (rgb2gray, remove_reflections, histogram_stretching,
                                            compress_grayscale, glcm_texture_features,
                                            save_glcm_matrix, visualize_features)

        if self.crop_file.exists():
            print("\n✅ Img_crop import completed!")
        else:
            raise FileNotFoundError("\n❌ Img_crop does not exist!")

        img_0 = cv2.imread(str(self.crop_file))
        img_gray = rgb2gray(img_0, self.dirs["img_gray"], self.dirs["img_gray"])
        print("\n✅ Gray-scale conversion completed!")
        gray_output = self.dirs["img_gray"] / "img_gray.jpg"
        self.assertTrue(gray_output.exists(), "\n❌ Img_gray not saved")
        print("\n✅ Img_gray saved!")
        self.assertEqual(len(img_gray.shape), 2, f"\n❌ Img_gray should be 2D, got {img_gray.shape}D")
        self.assertEqual(img_gray.dtype, np.uint8, f"\n❌ Img_gray should be uint8, got {img_gray.dtype}")
        print("\n✅ Img_gray meets the requirements of shape=2 and dtype=uint8!")
        print(f"✅✅ Test_2.1: gray-scale conversion verification passed!")

        img_repaired = remove_reflections(img_gray, self.dirs["binary_mask"], self.dirs["binary_mask"])
        print("\n✅ Extreme reflection removal completed!")
        src_dir = self.dirs["binary_mask"] / "img_repaired"
        shutil.move(str(src_dir / "binary_mask.jpg"), str(self.dirs["img_repaired"] / "img_repaired.jpg"))
        os.rmdir(src_dir)

        mask_output = self.dirs["binary_mask"] / "binary_mask.jpg"
        self.assertTrue(mask_output.exists(), "\n❌ Binary_mask not saved")
        print("\n✅ Binary_mask saved!")
        repaired_output = self.dirs["img_repaired"] / "img_repaired.jpg"
        self.assertTrue(repaired_output.exists(), "\n❌ Img_repaired not saved")
        print("\n✅ Img_repaired saved!")

        if repaired_output.exists():
            print("Minimum gray-scale value of img_repaired:", np.min(img_repaired))
            print("Maximum gray-scale value of img_repaired:", np.max(img_repaired))
            if mask_output.exists():
                mask_img = cv2.imread(str(mask_output), cv2.IMREAD_GRAYSCALE)
                self.assertEqual(mask_img.min(), 0, "\n❌ Binary_mask should be binary (0-255)")
                self.assertEqual(mask_img.max(), 255, "\n❌ Binary_mask should be binary (0-255)")
                print("Number of pixels with extreme reflection:", np.sum(mask_img == 255))
        print(f"✅✅ Test_2.2: extreme reflection removal verification passed!")

        img_256 = histogram_stretching(img_repaired, self.dirs["img_256"], self.dirs["img_256"])
        print("\n✅ Histogram_stretching completed!")
        stretch_output = self.dirs["img_256"] / "img_256.jpg"
        self.assertTrue(stretch_output.exists(), "\n❌ Img_256 not saved")
        print("\n✅ Img_256 saved!")
        self.assertEqual(img_256.dtype, np.float32, f"\n❌ Img_256 should be float32, got {img_256.dtype}")
        print("\n✅ Img_gray meets the requirements of shape=2 and dtype=uint8!")
        self.assertLess(int(np.min(img_256)),  2, f"\n❌ Minimum of img_256 should be 0 or 1, got {int(np.min(img_256))}!")
        self.assertLess(255 - int(np.max(img_256)), 2, f"\n❌ Maximum of img_256 should be 254 or 255, got 255 - {int(np.max(img_256))}!")
        print(f"✅✅ Test_2.3: Linear stretching verification passed!")

        img_comp = compress_grayscale(img_256, self.levels, self.dirs["img_comp"], self.dirs["img_comp"])
        print("\n✅ Gray-scale compression completed!")
        compress_output = self.dirs["img_comp"] / f"img_comp_{self.levels}.jpg"
        self.assertTrue(compress_output.exists(), "\n❌ Img_comp not saved!")
        print(f"\n✅ Img_comp_{self.levels} saved!")
        self.assertEqual(img_comp.dtype, np.uint8, f"\n❌ Img_comp should be uint8, got {img_comp.dtype}!")
        print("\n✅ Img_comp meets the requirement of dtype=uint8!")
        invalid_pixels = np.setdiff1d(img_comp, np.arange(0, 256, 256 // self.levels))
        self.assertEqual(len(invalid_pixels), 0,f"\n❌ Image_comp is not {self.levels} levels)!")
        print(f"\n✅ Image_comp meets the requirement of {self.levels} levels")
        print(f"✅✅ Test_2.4: Gray-scale compression verification passed!")

        features, glcm = glcm_texture_features(
            img_comp,
            properties=self.properties,
            distances=self.distances,
            angles=self.angles,
            levels=self.levels
        )
        print(f"\n✅ Features and glcm calculations completed!")

        self.assertEqual(len(features), len(self.properties), "\n❌ Several features lost!")
        self.assertEqual(list(features.keys()), self.properties, "\n❌ Features do not match!")
        print(f"\n✅ Features are completely matched!")
        for key, value in features.items():
            self.assertEqual(value.shape, (len(self.distances), len(self.angles)),
                             "\n❌ Features do not match!")
        print(f"\n✅ Feature values are completely matched!")
        print(f"✅✅ Test_2.5: GLCM_texture_features verification passed!")

        save_glcm_matrix(glcm, self.dirs["GLCM_matrix"], self.dirs["GLCM_matrix"])
        matrix_counts = len(list(Path(self.dirs["GLCM_matrix"]).glob("*.jpg")))
        self.assertEqual(matrix_counts, len(self.distances) * len(self.angles), "\n❌ Several GLCM matrices lost!")
        print(f"\n✅ All {matrix_counts} GLCM matrices saved!")
        print(f"✅✅ Test_2.6: Save_glcm_matrix verification passed!")

        visualize_features(features, self.distances, self.dirs["GLCM_feature_values"], self.dirs["GLCM_feature_values"])
        values_output = self.dirs["GLCM_feature_values"] / "GLCM_feature_values.xlsx"
        self.assertTrue(values_output.exists(), "\n❌ GLCM_feature_values not saved")
        print(f"\n✅ GLCM_feature_values saved!")

        wb = load_workbook(filename=values_output, read_only=True)
        sheet = wb.active
        num_features, num_angles, num_values = 0, 0, 0
        for row_idx, row in enumerate(sheet.iter_rows(values_only=True), 1):
            for col_idx, cell in enumerate(row, 1):
                if cell is not None and str(cell).strip() != '':
                    num_values += 1
                    if col_idx == 1:
                        num_features += 1
                    if row_idx == 1:
                        num_angles += 1
        wb.close()

        self.assertEqual(num_features, len(features), "\n❌ Features do not match!")
        self.assertEqual(num_angles, len(self.angles), "\n❌ Angles do not match!")
        self.assertEqual(num_values, len(features) + len(self.angles) + len(features) ** 2 * len(self.angles),
                         "\n❌ Values do not match!")
        print(f"\n✅ GLCM_feature_values are completely matched!")
        print(f"✅✅ Test_2.7: GLCM_feature_values verification passed!")
        print(f"✅✅ Test_2 (Test_2.1-Test_2.7) verification passed!")


class test_multi_distance_GLCM(unittest.TestCase):  # Test_3
    def setUp(self):
        self.root = Path(__file__).parent
        self.input_dir = self.root / "results" / "img_comp"
        self.input_file = self.input_dir / "img_comp_256.jpg"
        self.output_dir = self.root / "results" / "GLCM_matrix_data"
        self.output_dir.mkdir(parents=True, exist_ok=True)

        self.distances = list(range(1, 11))
        self.angles = [0, np.pi / 4, np.pi / 2, 3 * np.pi / 4]
        self.levels = 256

        if self.input_file.exists():
            print("\n✅ Image_comp_256 import completed!")
        else:
            raise FileNotFoundError("\n❌ Image_comp_256 does not exist!")

    def test_multi_distance_GLCM(self):
        from Multi_distance_GLCM import process_folder
        process_folder(str(self.input_dir), str(self.output_dir), self.distances, self.angles, self.levels)
        xlsx_counts = len(list(Path(self.output_dir).glob("*.xlsx")))
        self.assertEqual(xlsx_counts, max(self.distances), "\n❌ Several GLCM_matrice_data lost!")
        print(f"\n✅ All {xlsx_counts} GLCM_matrice_data saved!")

        excel_files = list(self.output_dir.glob("*.xlsx"))
        for file in excel_files:
            wb = load_workbook(filename=file, read_only=True)
            ws = wb.active
            wb.close()
            rows, cols = ws.max_row, ws.max_column
            self.assertEqual((rows, cols), (256, 256), f"\n❌ Shape of GLCM_matrice_data should be 256*256, got {rows, cols}!")
            print(f"Shape of {file} meets requirement of (256,256)")
        print("✅ All Shape of GLCM_matrice_data meets requirement of (256,256)")
        print(f"✅✅ Test_3 verification passed!")


if __name__ == '__main__':
    unittest.main()
